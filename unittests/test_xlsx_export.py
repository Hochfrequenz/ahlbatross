"""
Tests for Excel export functionality.
"""

from pathlib import Path
from typing import List, NamedTuple

import openpyxl  # type: ignore

from ahlbatross.enums.diff_types import DiffType
from ahlbatross.formats.xlsx import export_to_xlsx
from ahlbatross.models.ahb import AhbRow, AhbRowComparison, AhbRowDiff


class Formatversions(NamedTuple):
    """Test fixture for formatversion pairs."""

    previous_formatversion: str = "FV2504"
    subsequent_formatversion: str = "FV2510"


def test_xlsx_export_single_column_comparison(
    temp_excel_file: Path, ahb_row_comparison_single_column: List[AhbRowComparison]
) -> None:
    """
    Test output file exists.
    """
    export_to_xlsx(ahb_row_comparison_single_column, str(temp_excel_file))
    assert temp_excel_file.exists()
    assert temp_excel_file.stat().st_size > 0


def test_xlsx_export_multiple_column_comparison(
    temp_excel_file: Path, ahb_row_comparison_multiple_columns: List[AhbRowComparison]
) -> None:
    """Test export of complex data with all fields populated."""
    export_to_xlsx(ahb_row_comparison_multiple_columns, str(temp_excel_file))
    workbook = openpyxl.load_workbook(temp_excel_file)
    sheet = workbook.active

    # Check that all fields are present in the first data row
    row = 2  # First data row (after header)
    comp = ahb_row_comparison_multiple_columns[0]

    # Check previous version fields
    assert sheet.cell(row=row, column=1).value == comp.previous_formatversion.section_name
    assert sheet.cell(row=row, column=2).value == comp.previous_formatversion.segment_group_key
    assert sheet.cell(row=row, column=3).value == comp.previous_formatversion.segment_code
    assert sheet.cell(row=row, column=4).value == comp.previous_formatversion.data_element
    assert sheet.cell(row=row, column=5).value == comp.previous_formatversion.segment_id
    assert sheet.cell(row=row, column=6).value == comp.previous_formatversion.value_pool_entry
    assert sheet.cell(row=row, column=7).value == comp.previous_formatversion.name
    assert sheet.cell(row=row, column=8).value == comp.previous_formatversion.ahb_expression
    assert sheet.cell(row=row, column=9).value == comp.previous_formatversion.conditions


def test_xlsx_export_all_diff_types(
    temp_excel_file: Path, all_diff_types_ahb_row_comparisons: List[AhbRowComparison]
) -> None:
    """
    Test that all DIFF types are exported correctly.
    """
    export_to_xlsx(all_diff_types_ahb_row_comparisons, str(temp_excel_file))
    workbook = openpyxl.load_workbook(temp_excel_file)
    sheet = workbook.active

    # check correct number of rows: data + 1 (header)
    assert sheet.max_row == len(all_diff_types_ahb_row_comparisons) + 1

    diff_values = [sheet.cell(row=i, column=10).value or "" for i in range(2, sheet.max_row + 1)]
    expected_diffs = [comp.diff.diff_type.value for comp in all_diff_types_ahb_row_comparisons]
    assert diff_values == expected_diffs


def test_export_handles_none_values(temp_excel_file: Path, formatversions: Formatversions) -> None:
    """
    Test that None values are handled correctly.
    """
    comparisons = [
        AhbRowComparison(
            previous_formatversion=AhbRow(
                formatversion=formatversions.previous_formatversion,
                section_name=None,
                segment_group_key=None,
                segment_code=None,
                value_pool_entry=None,
                name=None,
            ),
            diff=AhbRowDiff(diff_type=DiffType.UNCHANGED),
            subsequent_formatversion=AhbRow(
                formatversion=formatversions.subsequent_formatversion,
                section_name=None,
                segment_group_key=None,
                segment_code=None,
                value_pool_entry=None,
                name=None,
            ),
        ),
    ]

    export_to_xlsx(comparisons, str(temp_excel_file))
    workbook = openpyxl.load_workbook(temp_excel_file)
    sheet = workbook.active

    # check None values are exported as empty strings
    for col in range(1, 10):
        assert sheet.cell(row=2, column=col).value == "" or sheet.cell(row=2, column=col).value is None


def test_export_header_formatting(
    temp_excel_file: Path, ahb_row_comparison_single_column: List[AhbRowComparison]
) -> None:
    """
    Test headers are formatted correctly.
    """
    export_to_xlsx(ahb_row_comparison_single_column, str(temp_excel_file))
    workbook = openpyxl.load_workbook(temp_excel_file)
    sheet = workbook.active

    assert sheet.max_row >= 1

    assert "Segmentname" in sheet.cell(row=1, column=1).value
    assert "Änderung" in sheet.cell(row=1, column=10).value


def test_xlsx_export_with_single_row(temp_excel_file: Path, basic_ahb_row: AhbRow) -> None:
    """
    Test export with a single row.
    """
    comparisons = [
        AhbRowComparison(
            previous_formatversion=basic_ahb_row,
            diff=AhbRowDiff(diff_type=DiffType.UNCHANGED),
            subsequent_formatversion=basic_ahb_row,
        )
    ]

    export_to_xlsx(comparisons, str(temp_excel_file))
    assert temp_excel_file.exists()
    assert temp_excel_file.stat().st_size > 0


def test_xlsx_export_large_dataset(temp_excel_file: Path, formatversions: Formatversions) -> None:
    """
    Test export with a large number of rows.
    """
    comparisons = []
    for i in range(1000):
        row = AhbRow(
            formatversion=formatversions.previous_formatversion,
            section_name=f"Section{i}",
            segment_group_key=f"SG{i}",
            segment_code=f"CODE{i}",
            value_pool_entry=None,
            name=None,
        )
        comparisons.append(
            AhbRowComparison(
                previous_formatversion=row,
                diff=AhbRowDiff(diff_type=DiffType.UNCHANGED),
                subsequent_formatversion=row,
            )
        )

    export_to_xlsx(comparisons, str(temp_excel_file))
    assert temp_excel_file.exists()
    assert temp_excel_file.stat().st_size > 0

    workbook = openpyxl.load_workbook(temp_excel_file)
    sheet = workbook.active
    assert sheet.max_row == len(comparisons) + 1  # data rows + 1 (header)
